"""
Statement Service

Handles account statement generation in multiple formats
"""

from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import Dict, Any, Optional
from datetime import date, datetime
from decimal import Decimal
import io

from backend.shared.database.deposit_models import (
    DepositAccount, DepositTransaction, DepositProduct
)
from backend.shared.database.models import Customer
from backend.shared.common.response import CustomException


class StatementService:
    """Service for generating account statements"""
    
    def __init__(self, db: Session, tenant_id: int, user_id: int):
        self.db = db
        self.tenant_id = tenant_id
        self.user_id = user_id
    
    def generate_statement(
        self,
        account_id: int,
        from_date: date,
        to_date: date
    ) -> Dict[str, Any]:
        """Generate statement data"""
        # Get account
        account = self._get_account(account_id)
        
        # Get customer
        customer = self.db.query(Customer).filter(
            Customer.id == account.customer_id
        ).first()
        
        # Get product
        product = self.db.query(DepositProduct).filter(
            DepositProduct.id == account.deposit_product_id
        ).first()
        
        # Get transactions
        transactions = self.db.query(DepositTransaction).filter(
            and_(
                DepositTransaction.deposit_account_id == account_id,
                DepositTransaction.transaction_date >= from_date,
                DepositTransaction.transaction_date <= to_date
            )
        ).order_by(DepositTransaction.transaction_date.asc()).all()
        
        # Get opening balance (balance before from_date)
        opening_txn = self.db.query(DepositTransaction).filter(
            and_(
                DepositTransaction.deposit_account_id == account_id,
                DepositTransaction.transaction_date < from_date
            )
        ).order_by(DepositTransaction.transaction_date.desc()).first()
        
        opening_balance = opening_txn.balance_after if opening_txn else account.principal_amount
        
        # Calculate totals
        total_deposits = sum(
            txn.amount for txn in transactions 
            if txn.transaction_type in ['deposit', 'interest_credit', 'opening', 'installment']
        )
        
        total_withdrawals = sum(
            txn.amount for txn in transactions 
            if txn.transaction_type in ['withdrawal', 'charge', 'penalty', 'interest_tds', 'closure']
        )
        
        total_interest = sum(
            txn.amount for txn in transactions 
            if txn.transaction_type == 'interest_credit'
        )
        
        closing_balance = transactions[-1].balance_after if transactions else opening_balance
        
        # Format transactions
        formatted_txns = []
        for txn in transactions:
            formatted_txns.append({
                "id": txn.id,
                "transaction_number": txn.transaction_number,
                "transaction_date": txn.transaction_date.isoformat(),
                "transaction_type": txn.transaction_type,
                "amount": float(txn.amount),
                "balance_before": float(txn.balance_before),
                "balance_after": float(txn.balance_after),
                "payment_mode": txn.payment_mode,
                "reference_number": txn.reference_number,
                "remarks": txn.remarks
            })
        
        return {
            "account": {
                "account_number": account.account_number,
                "account_type": account.account_type,
                "product_name": product.product_name if product else "Unknown",
                "opening_date": account.opening_date.isoformat(),
                "status": account.status
            },
            "customer": {
                "name": f"{customer.first_name} {customer.last_name}" if customer else "Unknown",
                "email": customer.email if customer else None,
                "phone": customer.phone if customer else None
            },
            "period": {
                "from_date": from_date.isoformat(),
                "to_date": to_date.isoformat()
            },
            "opening_balance": float(opening_balance),
            "closing_balance": float(closing_balance),
            "total_deposits": float(total_deposits),
            "total_withdrawals": float(total_withdrawals),
            "total_interest": float(total_interest),
            "transactions": formatted_txns
        }
    
    def generate_statement_pdf(
        self,
        account_id: int,
        from_date: date,
        to_date: date
    ) -> Dict[str, Any]:
        """Generate statement PDF"""
        statement = self.generate_statement(account_id, from_date, to_date)
        
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a365d'),
                spaceAfter=20,
                alignment=1
            )
            elements.append(Paragraph("ACCOUNT STATEMENT", title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Account & Customer Info
            info_data = [
                ["Account Number:", statement['account']['account_number']],
                ["Customer Name:", statement['customer']['name']],
                ["Account Type:", statement['account']['account_type'].upper()],
                ["Product:", statement['account']['product_name']],
                ["Statement Period:", f"{statement['period']['from_date']} to {statement['period']['to_date']}"],
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Summary
            summary_data = [
                ["Opening Balance:", f"₹{statement['opening_balance']:,.2f}"],
                ["Total Deposits:", f"₹{statement['total_deposits']:,.2f}"],
                ["Total Withdrawals:", f"₹{statement['total_withdrawals']:,.2f}"],
                ["Total Interest:", f"₹{statement['total_interest']:,.2f}"],
                ["Closing Balance:", f"₹{statement['closing_balance']:,.2f}"],
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f7fafc')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(summary_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Transactions
            if statement['transactions']:
                txn_header = ['Date', 'Type', 'Ref', 'Debit', 'Credit', 'Balance']
                txn_data = [txn_header]
                
                for txn in statement['transactions']:
                    debit = f"₹{txn['amount']:,.2f}" if txn['transaction_type'] in ['withdrawal', 'charge', 'penalty', 'closure'] else "-"
                    credit = f"₹{txn['amount']:,.2f}" if txn['transaction_type'] in ['deposit', 'interest_credit', 'opening', 'installment'] else "-"
                    
                    txn_data.append([
                        txn['transaction_date'],
                        txn['transaction_type'][:10],
                        (txn['reference_number'] or txn['transaction_number'][-8:])[:12],
                        debit,
                        credit,
                        f"₹{txn['balance_after']:,.2f}"
                    ])
                
                txn_table = Table(txn_data, colWidths=[0.9*inch, 1*inch, 1*inch, 1.1*inch, 1.1*inch, 1.2*inch])
                txn_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')])
                ]))
                
                elements.append(Paragraph("Transaction Details", styles['Heading2']))
                elements.append(Spacer(1, 0.1*inch))
                elements.append(txn_table)
            
            doc.build(elements)
            
            pdf_content = buffer.getvalue()
            buffer.close()
            
            filename = f"statement_{statement['account']['account_number']}_{from_date}_{to_date}.pdf"
            
            return {
                "pdf_content": pdf_content,
                "filename": filename
            }
            
        except ImportError:
            raise CustomException(
                status_code=500,
                message="PDF generation library not available"
            )
    
    def generate_statement_excel(
        self,
        account_id: int,
        from_date: date,
        to_date: date
    ) -> Dict[str, Any]:
        """Generate statement Excel"""
        statement = self.generate_statement(account_id, from_date, to_date)
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Statement"
            
            # Header
            ws['A1'] = "ACCOUNT STATEMENT"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:F1')
            
            # Account info
            row = 3
            ws[f'A{row}'] = "Account Number:"
            ws[f'B{row}'] = statement['account']['account_number']
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Customer Name:"
            ws[f'B{row}'] = statement['customer']['name']
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Period:"
            ws[f'B{row}'] = f"{statement['period']['from_date']} to {statement['period']['to_date']}"
            ws[f'A{row}'].font = Font(bold=True)
            
            # Summary
            row += 2
            ws[f'A{row}'] = "Opening Balance"
            ws[f'B{row}'] = statement['opening_balance']
            ws[f'A{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Closing Balance"
            ws[f'B{row}'] = statement['closing_balance']
            ws[f'A{row}'].font = Font(bold=True)
            
            # Transactions
            row += 2
            headers = ['Date', 'Type', 'Reference', 'Debit', 'Credit', 'Balance']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            for txn in statement['transactions']:
                row += 1
                ws.cell(row=row, column=1, value=txn['transaction_date'])
                ws.cell(row=row, column=2, value=txn['transaction_type'])
                ws.cell(row=row, column=3, value=txn['reference_number'] or txn['transaction_number'])
                
                if txn['transaction_type'] in ['withdrawal', 'charge', 'penalty', 'closure']:
                    ws.cell(row=row, column=4, value=txn['amount'])
                else:
                    ws.cell(row=row, column=5, value=txn['amount'])
                
                ws.cell(row=row, column=6, value=txn['balance_after'])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            
            filename = f"statement_{statement['account']['account_number']}_{from_date}_{to_date}.xlsx"
            
            return {
                "excel_content": excel_content,
                "filename": filename
            }
            
        except ImportError:
            raise CustomException(
                status_code=500,
                message="Excel generation library not available"
            )
    
    def email_statement(
        self,
        account_id: int,
        from_date: date,
        to_date: date,
        email_address: Optional[str] = None
    ) -> Dict[str, Any]:
        """Email statement to customer"""
        statement = self.generate_statement(account_id, from_date, to_date)
        pdf_result = self.generate_statement_pdf(account_id, from_date, to_date)
        
        # Get email address
        if not email_address:
            email_address = statement['customer']['email']
        
        if not email_address:
            raise CustomException(
                status_code=400,
                message="No email address available for customer"
            )
        
        # TODO: Integrate with email service
        # For now, return success with placeholder
        
        return {
            "email_address": email_address,
            "sent_date": datetime.utcnow().isoformat(),
            "filename": pdf_result['filename'],
            "status": "queued"  # In production, would be sent via email queue
        }
    
    def _get_account(self, account_id: int) -> DepositAccount:
        """Get and verify account"""
        account = self.db.query(DepositAccount).filter(
            and_(
                DepositAccount.id == account_id,
                DepositAccount.tenant_id == self.tenant_id,
                DepositAccount.is_deleted == False
            )
        ).first()
        
        if not account:
            raise CustomException(status_code=404, message="Account not found")
        
        return account
